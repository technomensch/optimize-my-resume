#!/usr/bin/env python3
"""
Convert requirements.md (job history creation - natural language) to Excel spreadsheet.
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def parse_requirements_md(filepath):
    """Parse the requirements markdown file and extract hierarchical data."""
    
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    rows = []
    lines = content.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Match Epic: # EPIC X.0.0.0: Title
        epic_match = re.match(r'^# EPIC (\d+\.0\.0\.0): (.+)$', line)
        if epic_match:
            epic_number = epic_match.group(1)
            epic_title = epic_match.group(2)
            # Look for description
            desc = ""
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if next_line.startswith('**Description:**'):
                    desc = next_line.replace('**Description:**', '').strip()
                    break
                elif next_line.startswith('#'):
                    break
                j += 1
            rows.append({
                'level': 1,
                'number': epic_number,
                'text': epic_title,
                'description': desc
            })
            i += 1
            continue
        
        # Match Feature: ## Feature X.Y.0.0: Title
        feature_match = re.match(r'^## Feature (\d+\.\d+\.0\.0): (.+)$', line)
        if feature_match:
            feature_number = feature_match.group(1)
            feature_title = feature_match.group(2)
            # Look for description
            desc = ""
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if next_line.startswith('**Description:**'):
                    desc = next_line.replace('**Description:**', '').strip()
                    break
                elif next_line.startswith('#'):
                    break
                j += 1
            rows.append({
                'level': 2,
                'number': feature_number,
                'text': feature_title,
                'description': desc
            })
            i += 1
            continue
        
        # Match User Story: ### User Story X.Y.Z.0: Title
        story_match = re.match(r'^### User Story (\d+\.\d+\.\d+\.0): (.+)$', line)
        if story_match:
            story_number = story_match.group(1)
            story_title = story_match.group(2)
            # Get the full user story text from next line
            desc = ""
            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if next_line.startswith('**As a**'):
                    desc = next_line
                    break
                elif next_line.startswith('#'):
                    break
                j += 1
            rows.append({
                'level': 3,
                'number': story_number,
                'text': story_title,
                'description': desc
            })
            i += 1
            continue
        
        # Match Acceptance Criteria: - **X.Y.Z.A.n:** Text
        ac_match = re.match(r'^- \*\*(\d+\.\d+\.\d+\.A\.\d+):\*\* (.+)$', line)
        if ac_match:
            ac_number = ac_match.group(1)
            ac_text = ac_match.group(2)
            rows.append({
                'level': 4,
                'number': ac_number,
                'text': ac_text,
                'description': ''
            })
            i += 1
            continue
        
        # Match Business Rule: - **X.Y.Z.R.n:** Text
        br_match = re.match(r'^- \*\*(\d+\.\d+\.\d+\.R\.\d+):\*\* (.+)$', line)
        if br_match:
            br_number = br_match.group(1)
            br_text = br_match.group(2)
            rows.append({
                'level': 5,
                'number': br_number,
                'text': br_text,
                'description': ''
            })
            i += 1
            continue
        
        i += 1
    
    return rows

def create_excel(rows, output_path):
    """Create Excel file with hierarchical structure."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    
    # Define column headers
    headers = [
        'Epic #', 'Feature #', 'Story #', 'AC/BR #', 'Rule Typ',
        'Epic Title', 'Feature Title', 'User Story Statement', 'Acceptance Criteria', 'Business Rule',
        'Functional Title / Description'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Style definitions
    epic_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    feature_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
    story_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ac_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    br_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    
    # Write data rows
    row_num = 2
    for item in rows:
        level = item['level']
        number = item['number']
        text = item['text']
        description = item.get('description', '')
        
        # Mapping mapping logic for best practice
        # Col 1-5: Numbers
        # Col 6-10: Primary Content
        # Col 11: Description/Meta
        
        if level == 1: # Epic
            ws.cell(row=row_num, column=1, value=number)
            ws.cell(row=row_num, column=6, value=text)
            ws.cell(row=row_num, column=11, value=description)
        elif level == 2: # Feature
            ws.cell(row=row_num, column=2, value=number)
            ws.cell(row=row_num, column=7, value=text)
            ws.cell(row=row_num, column=11, value=description)
        elif level == 3: # User Story
            ws.cell(row=row_num, column=3, value=number)
            # BEST PRACTICE: "User Story Statement" is the core content, "Title" is metadata
            ws.cell(row=row_num, column=8, value=description) # Statement
            ws.cell(row=row_num, column=11, value=text)        # Functional Title
        elif level == 4: # AC
            ws.cell(row=row_num, column=4, value=number)
            ws.cell(row=row_num, column=9, value=text)
        elif level == 5: # BR
            ws.cell(row=row_num, column=4, value=number) # Shared col with AC for numbering
            ws.cell(row=row_num, column=5, value='BUSRULE')
            ws.cell(row=row_num, column=10, value=text)
        
        # Apply fills based on level
        fill = None
        if level == 1:
            fill = epic_fill
        elif level == 2:
            fill = feature_fill
        elif level == 3:
            fill = story_fill
        elif level == 4:
            fill = ac_fill
        elif level == 5:
            fill = br_fill
        
        if fill:
            for col in range(1, 12):
                ws.cell(row=row_num, column=col).fill = fill
        
        # Apply borders
        for col in range(1, 12):
            ws.cell(row=row_num, column=col).border = thin_border
        
        row_num += 1
    
    # Set column widths
    column_widths = {
        'A': 14, 'B': 14, 'C': 16, 'D': 20, 'E': 18,
        'F': 40, 'G': 45, 'H': 55, 'I': 60, 'J': 60,
        'K': 80
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row heights for better readability
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    print(f"Total rows: {row_num - 1} (including header)")

if __name__ == '__main__':
    input_file = 'requirements.md'
    output_file = 'requirements.xlsx'
    
    print(f"Parsing: {input_file}")
    rows = parse_requirements_md(input_file)
    print(f"Found {len(rows)} items")
    
    # Print summary by level
    level_counts = {}
    for row in rows:
        level = row['level']
        level_counts[level] = level_counts.get(level, 0) + 1
    
    level_names = {1: 'Epics', 2: 'Features', 3: 'User Stories', 4: 'Acceptance Criteria', 5: 'Business Rules'}
    print("\nBreakdown:")
    for level in sorted(level_counts.keys()):
        print(f"  {level_names.get(level, f'Level {level}')}: {level_counts[level]}")
    
    create_excel(rows, output_file)
